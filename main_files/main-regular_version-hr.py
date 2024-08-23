################################## Yasmin's testing part:##################################

#                                         plan:

#                   splitting the dataset into 80% training, 20% test
#                   creating models repteatedly with the training set
#                   running Dana's code with the list frequency and intersection algorithm (Kmeans + "shallowing" data types + counting frequencies)
#                   doing the same with the 20% test data
#                   compare the results of the two and analyze the results


import sys
import os

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
sys.path.append("../TalentAI-research-new-last-update")
sys.path.append(
    "C:/Users/adler/OneDrive/Talent.AI/TalentAI-research-new-last-update/.venv/Lib/site-packages"
)
import logging

logging.basicConfig(filename="yasmin_error_log.txt", level=logging.ERROR)


from general_algos.Preprocess_for_hr import preProcess, KMeansClusterer
import csv
import numpy as np
import pandas as pd
import os as os
import openpyxl as openpyxl
from openpyxl import Workbook
from utilss import mean_generator


from statistic_regular_algo.Statistic_dot_product import Statistic_dot_product

# from statistic_regular_algo.Statistic_list_frequency import Statistic_list_frequency

from statistic_regular_algo.Statistic_intersection import Statistic_intersection

# from statistic_regular_algo.MixedDistance import MixedDistance
from sklearn.model_selection import train_test_split

# Save the reference to the original sys.stdout
original_stdout = sys.stdout

# Specify the file path where you want to redirect the prints
output_file_path = "../output.txt"

# Open the file in write mode, this will create the file if it doesn't exist
output_file = open(output_file_path, "w")

# Redirect sys.stdout to the file, comment this out if no need for print
# sys.stdout = output_file


import numpy as np
import re


types_list = [
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "numeric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "list",
    "categoric",
    "categoric",
    "categoric",
    "list",
    "list",
    "categoric",
    "categoric",
    "categoric",
    "numeric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "list",
    "categoric",
    "categoric",
    "categoric",
    "categoric",
    "numeric",
    "list",
    "list",
    "list",
]

i = 0
script_dir = os.path.dirname(
    os.path.abspath(__file__)
)  # Get directory where the script is located
csv_file_path = os.path.join(script_dir, "..", "datasets", "employes_flat_version.csv")

with open(csv_file_path, "r", encoding="utf-8") as csvfile:  # employes.csv
    # Create a CSV reader object
    csv_data = []
    csvreader = csv.reader(csvfile)
    i += 1
    # Iterate through each row in the CSV file

    for row in csvreader:
        # Append each row as a list to the csv_data list
        csv_data.append(row)

vectors = [np.array(f, dtype=object) for f in csv_data]
print("done initializing the vector with 100\% of the data")

df = pd.DataFrame(vectors)

# splitting the dataset into 80% training, 20% test:
# train_df, test_df = train_test_split(df, test_size=0.20, random_state=42)
train_vectors, test_vectors = train_test_split(vectors, test_size=0.20, random_state=42)

# ######################## training ########################

# train_vectors = train_df.values

hp, k = preProcess(train_vectors, types_list, Statistic_intersection, 9, 9)
# in order to run this you need to comment out the part that refers to one hot vector in kmeansclusterer

print("making model of dot for train data")
model = KMeansClusterer(
    num_means=k,
    distance=Statistic_intersection,
    repeats=2,
    type_of_fields=types_list,
    hyper_params=hp,
)

print("done initializing model with KMeansClusterer class")

# print(hp["list_freq_dict"])

model.cluster_vectorspace(train_vectors)

print("done making model")

# model.calc_min_max_dist(vectors)
model.get_wcss()
model.calc_distance_between_clusters()

train_results = model.getModelData()

# exit()

############################# test #################################


# test_vectors = test_df.values

hp, k = preProcess(test_vectors, types_list, Statistic_intersection, 9, 9)
# in order to run this you need to comment out the part that refers to one hot vector in kmeansclusterer

print("making model of dot for test data")
model = KMeansClusterer(
    num_means=k,
    distance=Statistic_intersection,
    repeats=2,  # was 8
    type_of_fields=types_list,
    hyper_params=hp,
)

# print(hp["list_freq_dict"])

model.cluster_vectorspace(test_vectors)

print("done making model")

# model.calc_min_max_dist(vectors)
model.get_wcss()
model.calc_distance_between_clusters()

test_results = model.getModelData()


################# Saving Results and Analysis #################

writer = pd.ExcelWriter('clustering_results.xlsx', engine='openpyxl')

# Convert results to DataFrame if they're not already, and write to Excel
# pd.DataFrame(train_results).to_excel(writer, sheet_name='Training Results')
# pd.DataFrame(test_results).to_excel(writer, sheet_name='Testing Results')

print(train_results, test_results)
writer._save()

def write_results_to_excel(filename, train_results, test_results):
    workbook = Workbook()
    
    # Write training results
    train_sheet = workbook.create_sheet("Training Results")
    for i, row in enumerate(train_results):
        for j, val in enumerate(row):
            cell = train_sheet.cell(row=i+1, column=j+1)
            cell.value = val
    
    # Write testing results
    test_sheet = workbook.create_sheet("Testing Results")
    for i, row in enumerate(test_results):
        for j, val in enumerate(row):
            cell = test_sheet.cell(row=i+1, column=j+1)
            cell.value = val
    
    # Save the workbook
    workbook.save(filename)
    print(f"Results saved successfully to {filename}")

# Call the function to save results
write_results_to_excel('clustering_results.xlsx', train_results, test_results)

# ################ Comparative Analysis #################

# try:
#     train_silhouette = model.get_Silhouette(train_vectors)
#     test_silhouette = model.get_Silhouette(test_vectors)
#     print(f"Training Silhouette Score: {train_silhouette}")
#     print(f"Testing Silhouette Score: {test_silhouette}")
# except Exception as e:
#     print("An error occurred while calculating silhouette scores:", str(e))

# # Closing the output file
# sys.stdout = original_stdout
# output_file.close()

exit()

##################################################################
print("###################### making model of Statistic_intersection ")
hp, k = preProcess(vectors, types_list, Statistic_intersection, 9, 9)
model = KMeansClusterer(
    num_means=k,
    distance=Statistic_intersection,
    repeats=5,
    type_of_fields=types_list,
    hyper_params=hp,
)
print(hp["frequencies"])
model.cluster_vectorspace(vectors)

print("done making model")

model.get_wcss()
model.calc_distance_between_clusters()
exit()
##########################################################3


# import sys
# import os

# sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
# sys.path.append("../TalentAI-research-new-last-update")
# sys.path.append(
#     "C:/Users/adler/OneDrive/Talent.AI/TalentAI-research-new-last-update/.venv/Lib/site-packages"
# )
# import logging

# logging.basicConfig(filename="yasmin_error_log.txt", level=logging.ERROR)


# from general_algos.Preprocess_for_hr import preProcess, KMeansClusterer
# import csv
# import numpy as np
# import pandas as pd
# import os as os
# import openpyxl as openpyxl
# from openpyxl import Workbook
# from utilss import mean_generator
# from sklearn.metrics import silhouette_score, silhouette_samples


# from statistic_regular_algo.Statistic_dot_product import Statistic_dot_product
# # from statistic_regular_algo.Statistic_list_frequency import Statistic_list_frequency

# from statistic_regular_algo.Statistic_intersection import Statistic_intersection
# # from statistic_regular_algo.MixedDistance import MixedDistance
# from sklearn.model_selection import train_test_split

# # Save the reference to the original sys.stdout
# original_stdout = sys.stdout

# # Specify the file path where you want to redirect the prints
# output_file_path = "../output.txt"

# # Open the file in write mode, this will create the file if it doesn't exist
# output_file = open(output_file_path, "w")

# # Redirect sys.stdout to the file, comment this out if no need for print
# # sys.stdout = output_file


# import numpy as np
# import re

# types_list = [
#     "categoric", "categoric", "categoric", "categoric", "numeric", "categoric",
#     "categoric", "categoric", "categoric", "categoric", "list", "categoric",
#     "categoric", "categoric", "list", "list", "categoric", "categoric",
#     "categoric", "numeric", "categoric", "categoric", "categoric", "categoric",
#     "categoric", "categoric", "categoric", "categoric", "categoric", "list",
#     "categoric", "categoric", "categoric", "categoric", "numeric", "list",
#     "list", "list"
# ]

# # Load dataset
# script_dir = os.path.dirname(os.path.abspath(__file__))
# csv_file_path = os.path.join(script_dir, "..", "datasets", "employes_flat_version.csv")

# with open(csv_file_path, "r", encoding="utf-8") as csvfile:
#     csv_data = list(csv.reader(csvfile))

# vectors = [np.array(f, dtype=object) for f in csv_data]
# print("done initializing the vector with 100% of the data")

# # Split the dataset into 80% training, 20% test
# train_vectors, test_vectors = train_test_split(vectors, test_size=0.20, random_state=42)

# # Function to write results to Excel
# def write_results_to_excel(filename, train_results, test_results):
#     workbook = Workbook()
#     train_sheet = workbook.create_sheet("Training Results", 0)
#     for i, row in enumerate(train_results):
#         for j, val in enumerate(row):
#             train_sheet.cell(row=i + 1, column=j + 1).value = val

#     test_sheet = workbook.create_sheet("Testing Results", 1)
#     for i, row in enumerate(test_results):
#         for j, val in enumerate(row):
#             test_sheet.cell(row=i + 1, column=j + 1).value = val

#     workbook.save(filename)
#     print(f"Results saved successfully to {filename}")

# # Function to process and cluster data
# def process_and_cluster(vectors, types_list, distance_function):
#     hp, k = preProcess(vectors, types_list, distance_function, 9, 9)
#     model = KMeansClusterer(
#         num_means=k,
#         distance=distance_function,
#         repeats=2,  # Reduce repeats to speed up
#         type_of_fields=types_list,
#         hyper_params=hp,
#     )
#     model.cluster_vectorspace(vectors)
#     results = model.getModelData()

#     # Calculate Silhouette Scores on a subset
#     subset_indices = np.random.choice(len(vectors), size=min(100, len(vectors)), replace=False)
#     subset_vectors = [vectors[i] for i in subset_indices]
#     labels = model.assign_labels(subset_vectors)
#     silhouette_avg = silhouette_score(subset_vectors, labels, metric=lambda x, y: distance_function(x, y, types_list, hp)[0])
#     sample_silhouette_values = silhouette_samples(subset_vectors, labels, metric=lambda x, y: distance_function(x, y, types_list, hp)[0])

#     return results, silhouette_avg, sample_silhouette_values.tolist()

# # Training
# train_results_intersection, train_silhouette_avg_intersection, train_sample_silhouette_values_intersection = process_and_cluster(train_vectors, types_list, Statistic_intersection)

# # Testing
# test_results_intersection, test_silhouette_avg_intersection, test_sample_silhouette_values_intersection = process_and_cluster(test_vectors, types_list, Statistic_intersection)

# # Convert results to list format for saving
# train_results_intersection_list = [[key, value] for key, value in train_results_intersection.items()]
# test_results_intersection_list = [[key, value] for key, value in test_results_intersection.items()]

# # Save results to Excel
# write_results_to_excel("clustering_results_intersection.xlsx", train_results_intersection_list, test_results_intersection_list)

# # Print key results to console for quick examination
# print(f"Training Silhouette Score (Intersection): {train_silhouette_avg_intersection}")
# print(f"Testing Silhouette Score (Intersection): {test_silhouette_avg_intersection}")

# # Export to a Jupyter notebook for further analysis
# import nbformat as nbf

# nb = nbf.v4.new_notebook()

# text = """
# ## Clustering Results Analysis

# This notebook analyzes the clustering results using the `Statistic_intersection` distance metric.
# """

# code = f"""
# import pandas as pd
# import matplotlib.pyplot as plt
# import seaborn as sns

# # Load clustering results
# train_results_intersection = {train_results_intersection}
# test_results_intersection = {test_results_intersection}

# # Load silhouette scores
# train_silhouette_avg_intersection = {train_silhouette_avg_intersection}
# train_sample_silhouette_values_intersection = {train_sample_silhouette_values_intersection}
# test_silhouette_avg_intersection = {test_silhouette_avg_intersection}
# test_sample_silhouette_values_intersection = {test_sample_silhouette_values_intersection}

# # Plot silhouette scores
# plt.figure(figsize=(10, 5))
# plt.title("Training Silhouette Scores (Intersection)")
# plt.hist(train_sample_silhouette_values_intersection, bins=20)
# plt.xlabel("Silhouette Score")
# plt.ylabel("Frequency")
# plt.show()

# plt.figure(figsize=(10, 5))
# plt.title("Testing Silhouette Scores (Intersection)")
# plt.hist(test_sample_silhouette_values_intersection, bins=20)
# plt.xlabel("Silhouette Score")
# plt.ylabel("Frequency")
# plt.show()

# # Visualize cluster distances
# train_cluster_distances_intersection = pd.DataFrame(train_results_intersection['cluster_distances'])
# test_cluster_distances_intersection = pd.DataFrame(test_results_intersection['cluster_distances'])

# plt.figure(figsize=(10, 8))
# sns.heatmap(train_cluster_distances_intersection, annot=True, cmap="coolwarm", cbar=True)
# plt.title("Training Cluster Distance Heatmap (Intersection)")
# plt.xlabel("Cluster")
# plt.ylabel("Cluster")
# plt.show()

# plt.figure(figsize=(10, 8))
# sns.heatmap(test_cluster_distances_intersection, annot=True, cmap="coolwarm", cbar=True)
# plt.title("Testing Cluster Distance Heatmap (Intersection)")
# plt.xlabel("Cluster")
# plt.ylabel("Cluster")
# plt.show()
# """

# nb['cells'] = [
#     nbf.v4.new_markdown_cell(text),
#     nbf.v4.new_code_cell(code)
# ]

# with open('clustering_analysis_intersection.ipynb', 'w') as f:
#     nbf.write(nb, f)

# print("Notebook saved as clustering_analysis_intersection.ipynb")

# # Closing the output file
# sys.stdout = original_stdout
# output_file.close()

# exit()
